import datetime

import xlwings as xw
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import holidays
from ColombianHolidays import days_before_holidays

# Establish a connection to a workbook
hours_report = xw.Book(r"C:\Users\15402\Desktop\python_hours_script\hours_report.xlsx")
schedule = xw.Book(r"C:\Users\15402\Desktop\python_hours_script\miner_schedule.xlsx")
wb = openpyxl.load_workbook(r"C:\Users\15402\Desktop\python_hours_script\miner_schedule.xlsx")

# Instantiate a sheet object
june_schedule = schedule.sheets['MUZO EMPLOYEES JUNE']
extra_horas = hours_report.sheets['HORA EXTRA']
domingo_festivo = hours_report.sheets['DOMINGO Y FESTIVO']
recargo_nocturno_sheet= hours_report.sheets['RECARGO NOCTURNO']

sheet = wb['MUZO EMPLOYEES JUNE']
lista = hours_report.sheets['LISTA']

#Colombian Holidays
colombianHolidays = holidays.Colombia(years = 2022)
days_before_holidays = days_before_holidays()
# variables for hora_en_que
nine_to_two = lista.range('H1').value
nine_to_twelve = lista.range('H2').value
twelve_to_two = lista.range('H3').value

# variables for tipo_de_recargo
nocturno = lista.range('G1').value
nocturno_dom_fest = lista.range('G2').value

def recargo_nocturno(row_start = 9,col_start = 'E',col_end = 'AH',name_col = 'D',cedula_col = 'C'):
    #initializing count and dictionary
    count = 0
    nocturno_dict = {}

    for i in range(row_start - 1, sheet.max_row + 1):  # looping down names
        name = name_col + str(i)
        cedula = cedula_col + str(i)

        name = june_schedule.range(name).value
        cedula = june_schedule.range(cedula).value

        for x in range(column_index_from_string(col_start), column_index_from_string(col_end) + 1, 1):  # looping through shifts
            shift = sheet.cell(row=i, column=x).value
            date = sheet.cell(row=row_start-2, column=x).value
            if shift != 'O' and name != 'NEW PERSON':  # the None clause gets rid of new miner
                if shift == 'N':
                    if date in colombianHolidays or (date.weekday() == 6): #nightshift on a sunday or holiday
                        if ((date + datetime.timedelta(days=1)) in colombianHolidays) or (date.weekday() + 1 == 6):#the next day is a holiday or sunday as well
                            dicts = {'name': name, 'cedula': cedula, 'date': date, 'shift': shift,
                                     'tipo_de_recargo': nocturno_dom_fest, 'hora_en_que': nine_to_two,
                                      'num_of_horas': 5}
                            count += 1
                            nocturno_dict[count] = dicts
                        else: # nightshift on a sunday or holiday with the next day not being a holiday or sunday
                            dicts = {'name': name, 'cedula': cedula, 'date': date, 'shift': shift,
                                     'tipo_de_recargo': nocturno_dom_fest, 'hora_en_que': nine_to_twelve,
                                     'num_of_horas': 3}
                            count += 1
                            nocturno_dict[count] = dicts

                            dicts = {'name': name, 'cedula': cedula, 'date': (date+datetime.timedelta(days = 1)), 'shift': shift,
                                     'tipo_de_recargo': nocturno, 'hora_en_que': twelve_to_two,
                                     'num_of_horas': 2}
                            count += 1
                            nocturno_dict[count] = dicts

                    elif ((date + datetime.timedelta(days=1)) in colombianHolidays) or (date.weekday()+1 == 6): #normal nightshift with the next day being a holiday or sunday
                        dicts = {'name': name, 'cedula': cedula, 'date': date, 'shift': shift,
                                 'tipo_de_recargo': nocturno, 'hora_en_que': nine_to_twelve,
                                 'num_of_horas': 3}
                        count += 1
                        nocturno_dict[count] = dicts

                        dicts = {'name': name, 'cedula': cedula, 'date': (date + datetime.timedelta(days=1)),
                                 'shift': shift,
                                 'tipo_de_recargo': nocturno_dom_fest, 'hora_en_que': twelve_to_two,
                                 'num_of_horas': 2}
                        count += 1
                        nocturno_dict[count] = dicts

                    else: #normal day night shift, next day not a sunday or a holiday
                        dicts = {'name': name, 'cedula': cedula, 'date': date, 'shift': shift,
                                 'tipo_de_recargo': nocturno, 'hora_en_que': nine_to_two,
                                 'num_of_horas': 5}
                        count += 1
                        nocturno_dict[count] = dicts
    return nocturno_dict

def recargo_nocturno_report(documento = 'C',nombre = 'D', tipo_de_recargo ='G', hora_en_que = 'I', date = 'H',
                           num_of_horas = 'J', starting_row = 5):
    nocturno_dict = recargo_nocturno()
    columnC = documento
    columnD = nombre
    columnG = tipo_de_recargo
    columnH = date
    columnI = hora_en_que
    columnJ = num_of_horas

    row = starting_row #starting row

    for entry in range(len(nocturno_dict)):
        cellC = columnC + str(entry+row)
        cellD = columnD + str(entry+row)
        cellG = columnG + str(entry+row)
        cellH = columnH + str(entry+row)
        cellI = columnI + str(entry+row)
        cellJ = columnJ + str(entry+row)

        recargo_nocturno_sheet.range(cellC).value = nocturno_dict[entry+1]['cedula']
        recargo_nocturno_sheet.range(cellD).value = nocturno_dict[entry+1]['name']
        recargo_nocturno_sheet.range(cellG).value = nocturno_dict[entry+1]['tipo_de_recargo']
        recargo_nocturno_sheet.range(cellI).value = nocturno_dict[entry+1]['hora_en_que']
        recargo_nocturno_sheet.range(cellH).value = nocturno_dict[entry+1]['date']
        recargo_nocturno_sheet.range(cellJ).value = nocturno_dict[entry+1]['num_of_horas']


recargo_nocturno_report()



