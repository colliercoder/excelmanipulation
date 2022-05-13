import pandas as pd
import openpyxl
import numpy as np


hours_report = openpyxl.load_workbook(r"C:\Users\15402\Desktop\python_hours_script\hours_report.xlsx")
schedule = openpyxl.load_workbook(r"C:\Users\15402\Desktop\python_hours_script\schedule.xlsx")

#pandas creating data frames from the sheets from the hours report

#df_domingo_festivo = pd.read_excel(r"C:\Users\15402\Desktop\python_hours_script\hours_report.xlsx", sheet_name="DOMINGO Y FESTIVO")
#df_hora_extra = pd.read_excel(r"C:\Users\15402\Desktop\python_hours_script\hours_report.xlsx", sheet_name="HORA EXTRA")
#df_recargo_nocturo = pd.read_excel(r"C:\Users\15402\Desktop\python_hours_script\hours_report.xlsx", sheet_name="RECARGO NOCTURO")

#pandas creating dataframes from the sheets from the schedule
#df_schedule_june = pd.read_excel(r"C:\Users\15402\Desktop\python_hours_script\schedule.xlsx", sheet_name="MUZO EMPLOYEES JUNE")

#printing the pandas dataframe
#print(df_schedule_june)


print(type(hours_report))
domingo_festivo = hours_report['DOMINGO Y FESTIVO']
hora_extra = hours_report['HORA EXTRA']
recargo_nocturo = hours_report['RECARGO NOCTURO']

schedule_june = schedule['MUZO EMPLOYEES JUNE']

#printing a title of a sheet
#print(schedule_june.title)

x = schedule_june['E8']
print('Row %s, Column %s is %s' % (x.row, x.column, x.value)) #prints Row 8, Column 5 is N

print('Cell %s is %s' % (x.coordinate, x.value)) # prints Cell E8 is N

print(x.value) #prints N

print(schedule_june['E8'].value) #prints N

print(schedule_june.max_row) #prints max row number
print(schedule_june.max_column) #prints max col number

#prints all the cedulas and names in the schedule for june
for i in range(8,24):
    print(i,schedule_june.cell(row=i, column = 3).value, schedule_june.cell(row=i, column = 4).value)

print(get_column_letter(1)) #print A
print(get_column_letter(schedule_june.max_column)) #print AL
print(column_index_from_string('T')) #print 20

print(tuple(schedule_june['C7':'AH23']))

for rowOfCellObjects in schedule_june['C7':'AH23']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)

#df = pd.read_excel(r"C:\Users\15402\Desktop\python_hours_script\schedule.xlsx",'MUZO EMPLOYEES JUNE',skiprows = 6,usecols = "C:AH",header)

print(df.head())