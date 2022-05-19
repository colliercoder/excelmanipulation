import xlwings as xw
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# Establish a connection to a workbook
hours_report = xw.Book(r"C:\Users\15402\Desktop\python_hours_script\hours_report.xlsx")
schedule = xw.Book(r"C:\Users\15402\Desktop\python_hours_script\miner_schedule.xlsx")
wb = openpyxl.load_workbook(r"C:\Users\15402\Desktop\python_hours_script\miner_schedule.xlsx")

# Instantiate a sheet object
june_schedule = schedule.sheets['MUZO EMPLOYEES JUNE']
extra_horas = hours_report.sheets['HORA EXTRA']
domingo_festivo = hours_report.sheets['DOMINGO Y FESTIVO']
recargo_nocturno = hours_report.sheets['RECARGO NOCTURNO']

sheet = wb['MUZO EMPLOYEES JUNE']
lista = hours_report.sheets['LISTA']

# Variables for novedad o tipo de hora 'G'
diurno = lista.range('C1').value
nocturno = lista.range('C2').value
domD = lista.range('C3').value
domN = lista.range('C4').value

# variables for turno normal del trabajador el el dia 'H'
day_turno_hrs = lista.range('B1').value
night_turno_hrs = lista.range('B2').value

# variables for turno en que se efectua la novedad 'I'
A_turno = lista.range('D1').value
C_turno = lista.range('D3').value
A_turno_domingo = lista.range('D4').value
C_turno_domingo = lista.range('D6').value

# variables for hora en que se efectua la novedad 'J'
day_start_hr = lista.range('B9').value
night_start_hr = lista.range('B10').value

# variables for actividad realiza and numero de horas
actividad = 'RAMPA JD'
num_hours = 2


def rangecreator(start_col='E', end_col='AH', names_column = 'D', cedula_column = 'C', start_row=9):
    # initializing variables
    count1 = 0
    count2 = 0
    count3 = 0
    bigdict = {}
    domingo_dict = {}
    nocturno_dict = {}

    for i in range(start_row - 1, sheet.max_row + 1):  # looping down names
        name = names_column + str(i)
        cedula = cedula_column + str(i)

        name = june_schedule.range(name).value
        cedula = june_schedule.range(cedula).value

        for x in range(column_index_from_string(start_col), column_index_from_string(end_col) + 1,
                       1):  # looping through shifts
            # print(sheet.cell(row=i, column=x).value)
            shift = sheet.cell(row=i, column=x).value
            date = sheet.cell(row=start_row - 2, column=x).value
            if shift != 'O' and name != 'NEW PERSON':  # the None clause gets rid of new miner
                dicts = {'name': name, 'cedula': cedula, 'date': date, 'shift': shift}
                count1 += 1
                bigdict[count1] = dicts

            if shift != 'O' and name != 'NEW PERSON' and date.weekday() == 6:
                doms = {'name': name, 'cedula': cedula, 'date': date, 'shift': shift}
                count2 += 1
                domingo_dict[count2] = doms

            if shift == 'N' and name != 'NEW PERSON':
                nocs = {'name': name, 'cedula': cedula, 'date': date, 'shift': shift}
                count3 += 1

                nocturno_dict[count3] = nocs

    return bigdict, domingo_dict, nocturno_dict