import xlwings as xw
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import holidays

# Establish a connection to a workbook
hours_report = xw.Book(r"C:\Users\15402\Desktop\python_hours_script\hours_report.xlsx")
schedule = xw.Book(r"C:\Users\15402\Desktop\python_hours_script\miner_schedule.xlsx")
wb = openpyxl.load_workbook(r"C:\Users\15402\Desktop\python_hours_script\miner_schedule.xlsx")

# Instantiate a sheet object
may_schedule = schedule.sheets['MUZO EMPLOYEES MAY']
extra_horas = hours_report.sheets['HORA EXTRA']
domingo_festivo = hours_report.sheets['DOMINGO Y FESTIVO']
recargo_nocturno = hours_report.sheets['RECARGO NOCTURNO']

sheet = wb['MUZO EMPLOYEES MAY']
lista = hours_report.sheets['LISTA']

#Colombian Holidays
colombianHolidays = holidays.Colombia()

# Variables for novedad o tipo de hora 'G'
diurno = lista.range('C1').value
nocturno = lista.range('C2').value
domD = lista.range('C3').value
domN = lista.range('C4').value

# variables for turno normal del trabajador el el dia 'H'
day_turno_hrs = lista.range('B1').value
night_turno_hrs = lista.range('B2').value

# variables for hora en que se efectua la novedad 'J'
day_start_hr = lista.range('B9').value
night_start_hr = lista.range('B10').value


def hora_extra_dict(row_start = 9,col_start = 'E',col_end = 'AI',name_col = 'D',cedula_col = 'C'):
    #initializing count and dictionary
    count = 0
    extraHrs_dict = {}

    for i in range(row_start - 1, sheet.max_row + 1):  # looping down names
        name = name_col + str(i)
        cedula = cedula_col + str(i)

        name = may_schedule.range(name).value
        cedula = may_schedule.range(cedula).value

        for x in range(column_index_from_string(col_start), column_index_from_string(col_end) + 1, 1):  # looping through shifts
            shift = sheet.cell(row=i, column=x).value
            date = sheet.cell(row=row_start-2, column=x).value
            if shift != 'O' and name != 'NEW PERSON':  # the None clause gets rid of new miner
                if shift == 'N':
                    if date.weekday() == 6 or date.weekday() in colombianHolidays:
                        dicts = {'name': name, 'cedula': cedula, 'date': date, 'shift': shift,
                                 'novedad': domN, 'turno normal': night_turno_hrs, 'hora_enque_efectua': night_start_hr,
                                 'actividad': 'RAMPA JD', 'num_of_horas': 2}
                    else:
                        dicts = {'name': name, 'cedula': cedula, 'date': date, 'shift': shift,
                                 'novedad': nocturno, 'turno normal': night_turno_hrs, 'hora_enque_efectua': night_start_hr,
                                 'actividad': 'RAMPA JD', 'num_of_horas': 2}
                if shift == 'D':
                    if date.weekday() == 6 or date.weekday() in colombianHolidays:
                        dicts = {'name': name, 'cedula': cedula, 'date': date, 'shift': shift,
                                 'novedad': domD, 'turno normal': day_turno_hrs, 'hora_enque_efectua': day_start_hr,
                                 'actividad': 'RAMPA JD', 'num_of_horas': 2}
                    else:
                        dicts = {'name': name, 'cedula': cedula, 'date': date, 'shift': shift,
                                 'novedad': diurno, 'turno normal': day_turno_hrs, 'hora_enque_efectua': day_start_hr,
                                 'actividad': 'RAMPA JD', 'num_of_horas': 2}

                count += 1
                extraHrs_dict[count] = dicts

    return extraHrs_dict

def hora_extra_report(documento = 'C',nombre = 'D', novedad =
                      'G', turno_normal = 'H', hora_se_efectua = 'I', actividad = 'J', date = 'K', num_of_horas = 'L',
                      starting_row = 5):
    extraHrs_dict = hora_extra_dict()
    columnC = documento
    columnD = nombre
    columnG = novedad
    columnH = turno_normal
    columnI = hora_se_efectua
    columnJ = actividad
    columnK = date
    columnL = num_of_horas

    row = starting_row #starting row

    for entry in range(len(extraHrs_dict)):
        cellC = columnC + str(entry+row)
        cellD = columnD + str(entry+row)
        cellG = columnG + str(entry+row)
        cellH = columnH + str(entry+row)
        cellI = columnI + str(entry+row)
        cellJ = columnJ + str(entry+row)
        cellK = columnK + str(entry+row)
        cellL = columnL + str(entry+row)

        extra_horas.range(cellC).value = extraHrs_dict[entry+1]['cedula']
        extra_horas.range(cellD).value = extraHrs_dict[entry+1]['name']
        extra_horas.range(cellG).value = extraHrs_dict[entry+1]['novedad']
        extra_horas.range(cellH).value = extraHrs_dict[entry+1]['turno normal']
        extra_horas.range(cellI).value = extraHrs_dict[entry+1]['hora_enque_efectua']
        extra_horas.range(cellJ).value = extraHrs_dict[entry+1]['actividad']
        extra_horas.range(cellK).value = extraHrs_dict[entry+1]['date']
        extra_horas.range(cellL).value = extraHrs_dict[entry+1]['num_of_horas']

hora_extra_report()