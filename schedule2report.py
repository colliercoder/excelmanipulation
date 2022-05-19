import xlwings as xw
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

#Establish a connection to a workbook
hours_report = xw.Book(r"C:\Users\15402\Desktop\python_hours_script\hours_report.xlsx")
schedule = xw.Book(r"C:\Users\15402\Desktop\python_hours_script\miner_schedule.xlsx")
wb = openpyxl.load_workbook(r"C:\Users\15402\Desktop\python_hours_script\miner_schedule.xlsx")

#Instantiate a sheet object
june_schedule = schedule.sheets['MUZO EMPLOYEES JUNE']
extra_horas = hours_report.sheets['HORA EXTRA']
domingo_festivo = hours_report.sheets['DOMINGO Y FESTIVO']
recargo_nocturno = hours_report.sheets['RECARGO NOCTURNO']


sheet = wb['MUZO EMPLOYEES JUNE']
lista = hours_report.sheets['LISTA']

#Variables for novedad o tipo de hora 'G'
diurno=lista.range('C1').value
nocturno=lista.range('C2').value
domD=lista.range('C3').value
domN=lista.range('C4').value

#variables for turno normal del trabajador el el dia 'H'
day_turno_hrs=lista.range('B1').value
night_turno_hrs=lista.range('B2').value

#variables for turno en que se efectua la novedad 'I'
A_turno=lista.range('D1').value
C_turno=lista.range('D3').value
A_turno_domingo=lista.range('D4').value
C_turno_domingo=lista.range('D6').value

#variables for hora en que se efectua la novedad 'J'
day_start_hr=lista.range('B9').value
night_start_hr=lista.range('B10').value

#variables for actividad realiza and numero de horas
actividad = 'RAMPA JD'
num_hours = 2
#

columnD = 'D' #names column
columnC = 'C' #cedula column

count = 0
count2 = 0
count3 = 0
bigdict = {}
domingo_dict = {}
nocturno_dict = {}

for i in range(8,sheet.max_row+1): #looping down names
    name = columnD + str(i)
    cedula = columnC + str(i)

    name = june_schedule.range(name).value
    cedula = june_schedule.range(cedula).value


    for x in range(column_index_from_string('E'), column_index_from_string('AH')+1, 1): #looping through shifts
        #print(sheet.cell(row=i, column=x).value)
        shift = sheet.cell(row=i, column=x).value
        date = sheet.cell(row=7, column=x).value
        if shift != 'O' and name != 'NEW PERSON': #the None clause gets rid of new miner
            dicts={'name':name,'cedula':cedula,'date':date,'shift':shift}
            count+=1

            bigdict[count] = dicts

        if shift != 'O' and name != 'NEW PERSON' and date.weekday() == 6:
            doms = {'name': name, 'cedula': cedula, 'date': date, 'shift': shift}
            count2 += 1

            domingo_dict[count2] = doms

        if shift == 'N' and name != 'NEW PERSON':
            nocs = {'name': name, 'cedula': cedula, 'date': date, 'shift': shift}
            count3 +=1

            nocturno_dict[count3] = nocs



columnC = 'C' #documento,----- cedula
columnD = 'D' #nombre apellido,----- name
columnG = 'G' # novedad/tipo de hora----- diurno,nocturno,domD,domN
columnH = 'H' #turno normal del trabajador el el dia ------ day_turno_hrs,night_turno_hrs
columnI = 'I' #turno en que se efectua la novedad------ A_turno, C_turno, A_turno_domingo, C_turno_domingo
columnJ = 'J' #hora en que se efectua la novedad ----------- day_start_hr, night_start_hr
columnK = 'K' #actividad realizada------actividad
columnL = 'L'
columnM = 'M' #numero de horas ---------  num_horas

row = 5 #starting row



for entry in range(len(bigdict)):
    cellC = columnC + str(entry+row)
    cellD = columnD + str(entry+row)
    cellG = columnG + str(entry+row)
    cellH = columnH + str(entry+row)
    cellI = columnI + str(entry+row)
    cellJ = columnJ + str(entry+row)
    cellK = columnK + str(entry+row)
    cellL = columnL + str(entry + row)
    cellM = columnM + str(entry+row)

    extra_horas.range(cellC).value = bigdict[entry+1]['cedula']
    extra_horas.range(cellD).value = bigdict[entry+1]['name']
    extra_horas.range(cellL).value = bigdict[entry + 1]['date']

    if bigdict[entry+1]['date'].weekday() == 6 and bigdict[entry+1]['shift'] == 'N':
        #HORA EXTRA SHEET
        extra_horas.range(cellG).value = domN
        extra_horas.range(cellH).value = night_turno_hrs
        extra_horas.range(cellI).value = C_turno_domingo
        extra_horas.range(cellJ).value = night_start_hr
        extra_horas.range(cellK).value = actividad
        extra_horas.range(cellM).value = num_hours
    elif bigdict[entry+1]['date'].weekday() == 6 and bigdict[entry+1]['shift'] == 'D':
        # HORA EXTRA SHEET
        extra_horas.range(cellG).value = domD
        extra_horas.range(cellH).value = day_turno_hrs
        extra_horas.range(cellI).value = A_turno_domingo
        extra_horas.range(cellJ).value = day_start_hr
        extra_horas.range(cellK).value = actividad
        extra_horas.range(cellM).value = num_hours
    elif bigdict[entry+1]['date'].weekday() != 6 and bigdict[entry+1]['shift'] == 'N':
        # HORA EXTRA SHEET
        extra_horas.range(cellG).value = nocturno
        extra_horas.range(cellH).value = night_turno_hrs
        extra_horas.range(cellI).value = C_turno
        extra_horas.range(cellJ).value = night_start_hr
        extra_horas.range(cellK).value = actividad
        extra_horas.range(cellM).value = num_hours
    else:
        # HORA EXTRA SHEET
        extra_horas.range(cellG).value = diurno
        extra_horas.range(cellH).value = day_turno_hrs
        extra_horas.range(cellI).value = A_turno
        extra_horas.range(cellJ).value = day_start_hr
        extra_horas.range(cellK).value = actividad
        extra_horas.range(cellM).value = num_hours


#DOMINGO Y FESTIVO
for entry in range(len(domingo_dict)):
    cellC = columnC + str(entry + row)
    cellD = columnD + str(entry + row)
    cellG = columnG + str(entry + row)
    cellH = columnH + str(entry + row)
    cellI = columnI + str(entry + row)
    cellJ = columnJ + str(entry + row)
    cellK = columnK + str(entry + row)
    cellL = columnL + str(entry + row)
    cellM = columnM + str(entry + row)

    if domingo_dict[entry + 1]['shift'] == 'N':
        # DOMINGO Y FESTIVO SHEET
        domingo_festivo.range(cellC).value = domingo_dict[entry + 1]['cedula']
        domingo_festivo.range(cellD).value = domingo_dict[entry + 1]['name']
        domingo_festivo.range(cellI).value = domingo_dict[entry + 1]['date']
        domingo_festivo.range(cellJ).value = 10
        domingo_festivo.range(cellG).value = 'DOMINGO'
        domingo_festivo.range(cellH).value = 'DOMINGO C'
    else:
        # DOMINGO Y FESTIVO SHEET
        domingo_festivo.range(cellC).value = domingo_dict[entry + 1]['cedula']
        domingo_festivo.range(cellD).value = domingo_dict[entry + 1]['name']
        domingo_festivo.range(cellI).value = domingo_dict[entry + 1]['date']
        domingo_festivo.range(cellJ).value = 10
        domingo_festivo.range(cellG).value = 'DOMINGO'
        domingo_festivo.range(cellH).value = 'DOMINGO A'


#RECARGO NOCTURNO

recargo_tiempo = '09:00PM'
recargo_horas = 7
for entry in range(len(nocturno_dict)):
    cellC = columnC + str(entry+row)
    cellD = columnD + str(entry+row)
    cellG = columnG + str(entry+row)
    cellH = columnH + str(entry+row)
    cellI = columnI + str(entry+row)
    cellJ = columnJ + str(entry+row)
    cellK = columnK + str(entry+row)
    cellL = columnL + str(entry + row)
    cellM = columnM + str(entry+row)

    if nocturno_dict[entry+1]['date'].weekday() == 6 and nocturno_dict[entry+1]['shift'] == 'N':
        #HORA EXTRA SHEET
        recargo_nocturno.range(cellC).value = nocturno_dict[entry + 1]['cedula']
        recargo_nocturno.range(cellD).value = nocturno_dict[entry + 1]['name']
        recargo_nocturno.range(cellH).value = nocturno_dict[entry + 1]['date']

        recargo_nocturno.range(cellG).value = 'NOCTURNO DOMINICAL O FESTIVO'

        recargo_nocturno.range(cellI).value = recargo_tiempo
        recargo_nocturno.range(cellJ).value = recargo_horas

    else:
        # HORA EXTRA SHEET
        recargo_nocturno.range(cellC).value = nocturno_dict[entry + 1]['cedula']
        recargo_nocturno.range(cellD).value = nocturno_dict[entry + 1]['name']
        recargo_nocturno.range(cellH).value = nocturno_dict[entry + 1]['date']

        recargo_nocturno.range(cellG).value = 'NOCTURNO'

        recargo_nocturno.range(cellI).value = recargo_tiempo
        recargo_nocturno.range(cellJ).value = recargo_horas

print("complete")





