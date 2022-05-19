import xlwings as xw
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import holidays
from ColombianHolidays import days_before_holidays

# Establish a connection to a workbook
hours_report = xw.Book(r"C:\Users\15402\Desktop\python_hours_script\hours_report.xlsx")
schedule = xw.Book(r"C:\Users\15402\Desktop\python_hours_script\miner_schedule.xlsx")
wb = openpyxl.load_workbook(r"C:\Users\15402\Desktop\python_hours_script\miner_schedule.xlsx")

# Instantiate a sheet object
june_schedule = schedule.sheets['MUZO EMPLOYEES JUNE']
extra_horas = hours_report.sheets['HORA EXTRA']
domingo_festivo = hours_report.sheets['DOMINGO Y FESTIVO']
recargo_nocturno = hours_report.sheets['RECARGO NOCTURNO']

sheet = wb['MUZO EMPLOYEES JUNE']
lista = hours_report.sheets['LISTA']

#Colombian Holidays
colombianHolidays = holidays.Colombia(years = 2022)
days_before_holidays = days_before_holidays()
# variables for novedad
domingo = lista.range('E1').value
festivo = lista.range('E2').value

# variables for turno en que se efectua
domingo_dia = lista.range('F1').value
festivo_dia = lista.range('F2').value
domingo_noche = lista.range('F3').value
festivo_noche = lista.range('F4').value
sabado_noche = lista.range('F5').value
noche_antes_festivo = lista.range('F6').value

#

def domingo_dict(row_start = 9,col_start = 'E',col_end = 'AH',name_col = 'D',cedula_col = 'C'):
    #initializing count and dictionary
    count = 0
    domingodict = {}

    for i in range(row_start - 1, sheet.max_row + 1):  # looping down names
        name = name_col + str(i)
        cedula = cedula_col + str(i)

        name = june_schedule.range(name).value
        cedula = june_schedule.range(cedula).value

        for x in range(column_index_from_string(col_start), column_index_from_string(col_end) + 1, 1):  # looping through shifts
            shift = sheet.cell(row=i, column=x).value
            date = sheet.cell(row=row_start-2, column=x).value
            if shift != 'O' and name != 'NEW PERSON':  # the None clause gets rid of new miner
                if shift == 'N':

                    if date in colombianHolidays: #Nightshift for a holiday
                        dicts = {'name': name, 'cedula': cedula, 'date': date, 'shift': shift,
                                 'novedad': festivo, 'turno_en_que': festivo_noche,
                                  'num_of_horas': 7}
                        count += 1
                        domingodict[count] = dicts

                    elif date.weekday() == 6: #Sunday night
                        dicts = {'name': name, 'cedula': cedula, 'date': date, 'shift': shift,
                                 'novedad': domingo, 'turno_en_que': domingo_noche,
                                 'num_of_horas': 7}
                        count += 1
                        domingodict[count] = dicts
                    elif date in days_before_holidays: #Nightshift before a holiday
                        dicts = {'name': name, 'cedula': cedula, 'date': date, 'shift': shift,
                                 'novedad': festivo, 'turno_en_que': noche_antes_festivo,
                                 'num_of_horas': 4}
                        count += 1
                        domingodict[count] = dicts
                    elif date.weekday() == 5:  # Saturday Night
                        dicts = {'name': name, 'cedula': cedula, 'date': date, 'shift': shift,
                                 'novedad': domingo, 'turno_en_que': sabado_noche,
                                 'num_of_horas': 4}
                        count += 1
                        domingodict[count] = dicts

                elif shift == 'D':

                    if date in colombianHolidays:
                        dicts = {'name': name, 'cedula': cedula, 'date': date, 'shift': shift,
                                 'novedad': festivo, 'turno_en_que': festivo_dia,
                                 'num_of_horas': 8}
                        count += 1
                        domingodict[count] = dicts

                    elif date.weekday() == 6:
                        dicts = {'name': name, 'cedula': cedula, 'date': date, 'shift': shift,
                                 'novedad': domingo, 'turno_en_que': domingo_dia,
                                 'num_of_horas': 8}
                        count += 1
                        domingodict[count] = dicts



    return domingodict

def domingo_festivo_report(documento = 'C',nombre = 'D', novedad ='G', turno_enque = 'H', date = 'I',
                           num_of_horas = 'J', starting_row = 5):
    domingodict = domingo_dict()
    columnC = documento
    columnD = nombre
    columnG = novedad
    columnH = turno_enque
    columnI = date
    columnJ = num_of_horas

    row = starting_row #starting row

    for entry in range(len(domingodict)):
        cellC = columnC + str(entry+row)
        cellD = columnD + str(entry+row)
        cellG = columnG + str(entry+row)
        cellH = columnH + str(entry+row)
        cellI = columnI + str(entry+row)
        cellJ = columnJ + str(entry+row)

        domingo_festivo.range(cellC).value = domingodict[entry+1]['cedula']
        domingo_festivo.range(cellD).value = domingodict[entry+1]['name']
        domingo_festivo.range(cellG).value = domingodict[entry+1]['novedad']
        domingo_festivo.range(cellH).value = domingodict[entry+1]['turno_en_que']
        domingo_festivo.range(cellI).value = domingodict[entry+1]['date']
        domingo_festivo.range(cellJ).value = domingodict[entry+1]['num_of_horas']


domingo_festivo_report()